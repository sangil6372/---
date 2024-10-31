import tkinter as tk
from tkinter import filedialog
import json
from openai import OpenAI
import re
import time
import shutil
import os
import pandas as pd
from openpyxl import load_workbook

# 처리 완료된 서브 폴더를 옮길 경로 설정
processed_folder_path = os.path.join(r'C:\Users\USER\Desktop\웅진_북센\booxen-refine-python\GPT/처리완료')  # '처리완료' 폴더 경로
os.makedirs(processed_folder_path, exist_ok=True)  # 폴더가 없을 경우 생성

from env import settings

# pm3
MODEL: str = 'gpt-4o'
client: OpenAI = OpenAI(api_key=settings.GPT_CONFIG['GPT_API_KEY'])


total_requests = 0

# 엑셀 파일에 데이터를 한 행씩 추가하는 함수
def append_to_excel(folder_name, request_tokens, output_tokens):
    # 엑셀 파일 경로 지정
    file_name = os.path.join(processed_folder_path, 'gpt_tokens.xlsx')

    # 엑셀 파일이 존재하는지 확인
    if not os.path.exists(file_name):
        # 파일이 없으면 새로운 파일을 생성
        df = pd.DataFrame(columns=['Folder Name', 'Request Tokens', 'Output Tokens'])
        df.to_excel(file_name, index=False)

    # 기존 엑셀 파일 불러오기
    book = load_workbook(file_name)
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    writer.book = book

    # 새 데이터 추가
    new_data = pd.DataFrame([[folder_name, request_tokens, output_tokens]],
                            columns=['Folder Name', 'Request Tokens', 'Output Tokens'])

    # 기존 시트에 데이터 추가
    new_data.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)

    # 엑셀 파일 저장
    writer.save()

# 파일 탐색기를 사용하여 디렉토리를 선택하는 함수
def select_directory(dialog_title):
    root = tk.Tk()
    root.withdraw()  # Tkinter 윈도우를 숨김
    directory_path = filedialog.askdirectory(title=dialog_title)
    return directory_path

# 텍스트 박스 내부에 수식 중심이 포함되어 있는지 확인하는 함수
def is_inline_formula(shape, text_bboxes):
    x1, y1 = shape["points"][0]
    x2, y2 = shape["points"][1]
    cx = (x1 + x2) / 2  # 수식의 x 중심 좌표
    cy = (y1 + y2) / 2  # 수식의 y 중심 좌표

    # 수식 중심이 텍스트 박스 내에 포함되는지 확인
    for tx1, ty1, tx2, ty2 in text_bboxes:
        if tx1 <= cx <= tx2 and ty1 <= cy <= ty2:
            return True
    return False

# 두 박스의 y 좌표가 겹치는지 확인하는 함수
def is_overlapping_y(box1, box2):
    y1_min = min(box1["points"][0][1], box1["points"][1][1])
    y1_max = max(box1["points"][0][1], box1["points"][1][1])

    y2_min = min(box2["points"][0][1], box2["points"][1][1])
    y2_max = max(box2["points"][0][1], box2["points"][1][1])

    # 겹치는 부분의 시작과 끝
    overlap_start = max(y1_min, y2_min)
    overlap_end = min(y1_max, y2_max)

    # 겹치는 길이
    overlap_length = max(0, overlap_end - overlap_start)

    # 각 박스의 높이
    height1 = y1_max - y1_min
    height2 = y2_max - y2_min

    # 겹치는 길이가 두 박스 높이의 50% 이상인지 확인
    if (overlap_length >= 0.5 * height1) or (overlap_length >= 0.5 * height2):
        return True
    return False


# 커스텀 정렬 함수 정의
def custom_sort(shapes):
    sorted_shapes = []

    while shapes:
        current = shapes.pop(0)
        overlaps = [current]
        non_overlaps = []
        # 현재 상태를 디버깅하기 위해 출력

        # 겹치는 박스를 찾기
        for other in shapes:
            if current['label'] == other['label'] and is_overlapping_y(current, other):
                overlaps.append(other)
            else:
                non_overlaps.append(other)

        # 겹치는 박스들을 x 좌표로 정렬
        overlaps_sorted = sorted(overlaps, key=lambda item: item["points"][0][0])
        sorted_shapes.extend(overlaps_sorted)

        # 남은 박스들을 다음 순서로 처리
        shapes = non_overlaps

    return sorted_shapes


# 이미지 파일을 base64로 인코딩하는 함수
def encode_image(image_path):
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode('utf-8')
    return encoded_string


import io
import base64
from PIL import Image
import os


def crop_image(image_path, coordinates, block_index, page_number):
    try:
        img = Image.open(image_path)
        img_width, img_height = img.size

        # 좌표를 픽셀 단위로 변환
        x1 = int(coordinates["x1"] * img_width)
        y1 = int(coordinates["y1"] * img_height)
        x2 = int(coordinates["x2"] * img_width)
        y2 = int(coordinates["y2"] * img_height)

        # 좌표 정렬 (잘못된 순서라면 교정)
        x1, x2 = min(x1, x2), max(x1, x2)
        y1, y2 = min(y1, y2), max(y1, y2)

        # 좌표가 이미지 경계를 벗어나면 None 리턴
        if x1 < 0 or y1 < 0 or x2 > img_width or y2 > img_height:
            print(f"Block {block_index} on page {page_number} is out of bounds. Skipping this block.")
            return None

        # 크롭할 영역이 유효한지 확인 (너비 또는 높이가 0인 경우 처리)
        if x2 <= x1 or y2 <= y1:
            print(f"Invalid crop area for block {block_index} on page {page_number}. Skipping this block.")
            return None

        # 이미지를 잘라냄
        cropped_img = img.crop((x1, y1, x2, y2))

        # 너비와 높이에 따라 크기 조정
        if cropped_img.width < 800:
            max_width = 512
        else:
            max_width = 1024

        if cropped_img.height < 800:
            max_height = 512
        else:
            max_height = 1024

        # 비율을 유지하며 크기 조정
        cropped_img.thumbnail((max_width, max_height))

        # 메모리 버퍼에 이미지 저장
        buffered = io.BytesIO()
        cropped_img.save(buffered, format="PNG")

        # base64로 인코딩
        encoded_cropped_img = base64.b64encode(buffered.getvalue()).decode('utf-8')

        return encoded_cropped_img

    except Exception as e:
        print(f"Error cropping image for block {block_index} on page {page_number}: {e}")
        return None


# 이미지 파일의 해상도를 JSON 데이터에서 0일 경우 대체하는 함수
def get_image_resolution_if_needed(image_path, extracted_data):
    try:
        if extracted_data["imageWidth"] == 0 or extracted_data["imageHeight"] == 0:
            img = Image.open(image_path)
            img_width, img_height = img.size

            # JSON 데이터에 이미지 해상도가 없을 경우 대체
            if extracted_data["imageWidth"] == 0:
                extracted_data["imageWidth"] = img_width
            if extracted_data["imageHeight"] == 0:
                extracted_data["imageHeight"] = img_height

    except Exception as e:
        print(f"Error getting image resolution for {image_path}: {e}")


gpt_failure_keywords = [
    "extracted text", "extract text", "sorry"
    "I'm unable to assist",
    "I'm unable to verify the extracted text", "I'm unable to extract text directly from an image",
    "I can't help with that", "I'm unable to access the images",
    "I'm unable to extract", "text extraction",
    "이미지에서 추출", "수정된 텍스트",
    "이미지를 분석",
    "추출된 텍스트", "텍스트를 추출", "죄송"
]



# GPT API를 사용한 오타 검증 함수
def check_image_and_text_with_gpt(image_base64, text):
    global total_requests
    total_requests += 1  # Increment the request counter

    try:

        response = client.chat.completions.create(
            model=MODEL,
            messages=[
                {
                    "role": "system",
                    "content": "당신은 이미지의 수식을 LaTeX코드로 변환하는 도우미입니다."
                },
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": r"첨부한 이미지는 도서에서 수식박스를 크롭한 작은 이미지입니다."
                                    r"해당 수식을 LaTeX코드로 변환해주세요."
                                    r"package, 포맷팅이나 \(\), \[,\]등의 수식 환경, 태그 등의 코드 블록 표시나 포맷팅 텍스트는 제외 해주세요."
                                    r"없는 내용을 추가하거나 수식을 잘못 해석하는 경우가 없도록 이미지의 내용을 객관적으로 그대로 반영해 주세요."
                                    r"출력 시 enter를 사용하지 않고 한 줄로 길게 출력 해주세요. "
                                    r"LaTeX 명령어 사용 할 때 빈칸이나 중괄호를 사용해서 인수와 구분해줘"
                                    r"수식 외의 부가 설명은 전부 제외하고 변환된 LaTeX코드만 그대로 출력해 주세요."
                        },
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{image_base64}"
                            }
                        }
                    ]
                }
            ],
            max_tokens=3000
        )

        corrected_text = response.choices[0].message.content.strip()

        # 만약 correct_text가 비어있을 경우, gpt 답변이 불가능할 경우(자주 나오는 패턴)
        if corrected_text is None:
            # None일 경우
            print(f"[GPT_ERROR] : GPT_OUTPUT: None, ORIGINAL_TEXT: {text}")
            corrected_text = text  # corrected_text를 text로 대체
        elif corrected_text == "" or corrected_text.isspace():
            # 비어있거나 공백인 경우
            print(f"[GPT_ERROR] : GPT_OUTPUT: Empty or Whitespace, ORIGINAL_TEXT: {text}")
            corrected_text = text  # corrected_text를 text로 대체
        elif any(keyword.lower() in corrected_text.lower() for keyword in gpt_failure_keywords):
            # gpt_failure_keywords 중 하나라도 포함된 경우
            print(f"[GPT_ERROR] : GPT_OUTPUT: {corrected_text}, ORIGINAL_TEXT: {text}")
            corrected_text = text  # corrected_text를 text로 대체

        # 사용된 토큰 수 추출 및 반환
        prompt_tokens = response.usage.prompt_tokens
        completion_tokens = response.usage.completion_tokens
        total_used_tokens = response.usage.total_tokens

        return corrected_text, prompt_tokens, completion_tokens, total_used_tokens

    except Exception as e:
        print(f"An error occurred during API call: {e}")
        return None, 0, 0, 0


# 페이지 번호를 파일 이름에서 추출하는 함수
def extract_page_number(file_name):
    match = re.search(r'_(\d+)\.', file_name)
    return int(match.group(1)) if match else None


# 파일 이름에서 언더바 기준으로 마지막 부분(쪽수) 추출
def extract_page_number_from_filename(filename):
    return filename.rsplit('_', 1)[-1].replace('.json', '').replace('.png', '')


# 파일 이름에서 첫 번째 언더바 앞에 있는 숫자가 5개 미만일 경우 제거하는 함수
def clean_filename(filename):
    parts = filename.split('_', 1)  # 언더바를 기준으로 나눔
    if parts[0].isdigit() and len(parts[0]) < 5:
        return '_'.join(parts[1:])  # 첫 번째 언더바 앞 숫자가 5개 미만이면 제거하고 나머지 반환
    return filename  # 그렇지 않으면 원본 반환


# 디렉토리 선택
json_folder_path = select_directory("OCR 텍스트가 저장된 JSON 파일들이 있는 폴더를 선택하세요")

# 설정한 경로에 gpt_updated/ 폴더 생성
gpt_updated_folder_path = os.path.join(r"C:\Users\USER\Desktop\gpt_test_result", 'gpt_updated')
os.makedirs(gpt_updated_folder_path, exist_ok=True)

# 전체 폴더에서 사용된 토큰을 저장할 변수들
total_prompt_tokens_all_files = 0
total_completion_tokens_all_files = 0
total_tokens_all_files = 0

start_time = time.time()  # 시작 시간 기록
# 하위 폴더들을 순회하면서 각각의 폴더 내에 json과 png 파일을 처리
for subfolder_name in os.listdir(json_folder_path):
    subfolder_path = os.path.join(json_folder_path, subfolder_name)

    # 각 서브폴더에서 사용된 토큰을 전체 합산
    subfolder_prompt_tokens_all_files = 0
    subfolder_completion_tokens_all_files = 0



    if os.path.isdir(subfolder_path):  # 하위 폴더인지 확인
        json_files = [f for f in os.listdir(subfolder_path) if f.endswith('.json')]
        png_files = [f for f in os.listdir(subfolder_path) if f.endswith('.png')]

        print(f"Processing subfolder: {subfolder_name}")
        # png 파일이 없는 경우 해당 폴더 스킵
        if len(png_files) < 1:
            print(f"{subfolder_name}에는 png 파일이 하나도 없습니다.")
            continue

        # json png 이름 비교 다를 경우 continue
        # 첫 번째 JSON 파일과 첫 번째 PNG 파일을 골라서 비교
        first_json_file = json_files[0]
        first_png_file = png_files[0]

        # 파일 이름에서 첫 번째 언더바 앞의 숫자가 5개 미만이면 제거
        cleaned_json_filename = clean_filename(first_json_file.replace('.json', ''))
        cleaned_png_filename = clean_filename(first_png_file.replace('.png', ''))

        # 비교 후 다르면 continue
        if cleaned_json_filename.split('_')[0] != cleaned_png_filename.split('_')[0]:
            print(f"파일 이름이 다릅니다: {cleaned_json_filename} != {cleaned_png_filename}")
            continue

        # gpt_updated/ 안에 하위 폴더 이름과 동일한 폴더 생성
        updated_subfolder_path = os.path.join(gpt_updated_folder_path, subfolder_name)
        os.makedirs(updated_subfolder_path, exist_ok=True)


        # JSON 파일 처리
        for json_filename in json_files:  # 중복된 os.listdir 제거
            json_file_path = os.path.join(subfolder_path, json_filename)

            # output 폴더에 업데이트된 해당 json 파일이 존재할 경우 continue
            output_json_file_path = os.path.join(updated_subfolder_path, json_filename)
            if os.path.exists(output_json_file_path):
                print(f"이미 업데이트된 파일이 존재합니다: {json_filename}. 건너뜁니다.")
                continue

            # JSON 파일 읽기
            with open(json_file_path, 'r', encoding='utf-8') as json_file:
                extracted_data = json.load(json_file)

            # json 파일 이름에서 페이지 번호 추출
            json_page_number = extract_page_number_from_filename(json_filename)
            if json_page_number is None:
                print(f"페이지 번호를 추출할 수 없습니다: {json_filename}")
                continue

            image_file_name = None
            for png_filename in png_files:
                png_page_number = extract_page_number_from_filename(png_filename)
                if json_page_number == png_page_number:
                    image_file_name = png_filename
                    break

            if not image_file_name:  # error 출력 필요
                print(f"PNG 파일을 찾을 수 없습니다: {json_filename}")
                continue

            # 이미지 파일 경로 생성
            #           image_file_name = f"{os.path.splitext(json_filename)[0]}.png"  # JSON 파일과 같은 이름의 이미지 파일
            image_path = os.path.join(subfolder_path, image_file_name)

            # 이미지 폴더 순회 및 오타 검증
            try:
                block_index = 0  # 블록 인덱스 초기화
                page_prompt_tokens = 0
                page_completion_tokens = 0
                page_total_tokens = 0

                # JSON 파일의 해상도 값이 0일 경우 이미지 파일에서 대체하는 함수 호출
                get_image_resolution_if_needed(image_path, extracted_data)

                # 그 전에 일단 정렬
                extracted_data["shapes"] = custom_sort(extracted_data["shapes"])

                # shapes 리스트에서 'inline formula'만 처리
                for shape in extracted_data["shapes"]:
                    # 텍스트 박스 좌표를 추출하여 text_bboxes에 저장
                    text_bboxes = [
                        (text_shape["points"][0][0], text_shape["points"][0][1],
                         text_shape["points"][1][0], text_shape["points"][1][1])
                        for text_shape in extracted_data["shapes"]
                        if text_shape["label"] in ["TEXT", "FOOTNOTE", "REFERENCE"]  # 수정된 부분
                    ]

                    # 'inline formula'만 처리
                    if shape["label"] == "FORMULA" and is_inline_formula(shape, text_bboxes):
                        text = shape["latex"]
                        coordinates = {
                            "x1": shape["points"][0][0] / extracted_data["imageWidth"],
                            "y1": shape["points"][0][1] / extracted_data["imageHeight"],
                            "x2": shape["points"][1][0] / extracted_data["imageWidth"],
                            "y2": shape["points"][1][1] / extracted_data["imageHeight"]
                        }

                        # 텍스트 블록별로 이미지 크롭
                        encoded_cropped_img = crop_image(image_path, coordinates, block_index, json_page_number)

                        if not encoded_cropped_img:
                            block_index += 1
                            continue

                        # GPT API로 크롭된 이미지와 텍스트 검증
                        corrected_text, prompt_tokens, completion_tokens, total_used_tokens = check_image_and_text_with_gpt(
                            encoded_cropped_img, text
                        )

                        # 페이지별 토큰 수 합산
                        page_prompt_tokens += prompt_tokens
                        page_completion_tokens += completion_tokens
                        page_total_tokens += total_used_tokens

                        # 결과를 업데이트
                        if corrected_text:
                            shape["latex"] = corrected_text

                        block_index += 1  # 블록 인덱스 증가

                subfolder_prompt_tokens_all_files += page_prompt_tokens
                subfolder_completion_tokens_all_files += page_completion_tokens

                # 폴더 이름, 요청 토큰, 아웃풋 토큰

                # 각 페이지에서 사용된 토큰을 전체 합산
                total_prompt_tokens_all_files += page_prompt_tokens
                total_completion_tokens_all_files += page_completion_tokens

            except Exception as e:
                print(f"An error occurred in file {json_filename}: {e}", flush=True)
                continue

            # 수정된 데이터를 새 JSON 파일로 저장 (updated_subfolder_path 폴더에 저장)
            output_json_file_path = os.path.join(updated_subfolder_path, json_filename)
            with open(output_json_file_path, 'w', encoding='utf-8') as output_json_file:
                json.dump(extracted_data, output_json_file, ensure_ascii=False, indent=4)
            print(f"오타 검증이 완료되었으며, 결과가 {output_json_file_path} 파일에 저장되었습니다.", flush=True)

        # JSON과 PNG 파일의 앞부분 숫자 여부 확인
        json_parts = first_json_file.split('_')
        png_parts = first_png_file.split('_')

        json_has_number = json_parts[0].isdigit() and len(json_parts[0]) < 5
        png_has_number = png_parts[0].isdigit() and len(png_parts[0]) < 5

        # JSON 파일에 번호가 있고, PNG 파일에 번호가 없는 경우 PNG 파일에도 번호 추가
        for png_filename in png_files:
            png_file_path = os.path.join(subfolder_path, png_filename)
            output_png_file_path = os.path.join(updated_subfolder_path, png_filename)

            # JSON 파일에 번호가 있고, PNG 파일에 번호가 없는 경우 PNG 파일에도 번호 추가
            if json_has_number and not png_has_number:
                new_png_filename = f"{json_parts[0]}_{png_filename}"  # JSON 파일 앞 번호를 붙여서 PNG 파일 이름 변경
            # JSON 파일에 번호가 없고, PNG 파일에 번호가 있는 경우 PNG 파일에서 번호 제거
            elif not json_has_number and png_has_number:
                new_png_filename = '_'.join(png_parts[1:])  # PNG 파일 앞 번호 제거
            else:
                new_png_filename = png_filename  # 둘 다 번호가 있거나 없으면 그대로 유지

            # 새 파일 경로 생성 및 파일 이동
            output_png_file_path = os.path.join(updated_subfolder_path, new_png_filename)
            try:
                os.rename(png_file_path, output_png_file_path)
            except:
                print("error : 이미 옮겨졌으므로 건너 뜁니다.")
                continue

            print(f"PNG 파일이 {output_png_file_path}로 이동되었습니다.")

end_time = time.time()  # 종료 시간 기록
elapsed_time = end_time - start_time  # 소요 시간 계산

print(f"총 소요 시간 : {elapsed_time:.2f} seconds")
print(f"요청 횟수: {total_requests}")
# 전체 폴더에서 사용된 총 토큰 출력
print(f"요청 토큰 수: {total_prompt_tokens_all_files}")
print(f"응답 토큰 수: {total_completion_tokens_all_files}")
print(f"총 토큰 수: {total_tokens_all_files}")
